from openpyxl import Workbook, load_workbook

# Future features ...
def create_excel(self, excel_file):
    """ Crea un archivo Excel a partir de los datos en la instancia """
    wb = Workbook()
    sheet = wb.active
    for row in self.__table:
        sheet.append(row)
    wb.save(excel_file)

def write_to_existing_excel(self, excel_file, base_excel_file):
    """ Escribe en un archivo Excel existente a partir de los datos en la instancia """
    # Abrir el archivo Excel base
    wb_base = load_workbook(filename=base_excel_file)
    # Acceder a la hoja existente en el archivo base
    sheet_base = wb_base.active
    # Empezar desde la primera fila
    first_empty_row = 1  
    # Iterar sobre los datos de la instancia y escribir en la hoja existente
    for row in self.__table:
        for idx, value in enumerate(row, start=1):
            sheet_base.cell(row=first_empty_row, column=idx, value=value)
        first_empty_row += 1
    # Guardar los cambios en el archivo Excel existente
    wb_base.save(excel_file)